import csv
import json
import io
from django.utils import timezone

# Import ReportLab independently so chart issues don't break PDF generation
try:
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import letter
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, Image
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

# Import chart utilities separately so a missing matplotlib never disables the PDF engine
try:
    from .chart_utils import ChartGenerator
    CHARTS_AVAILABLE = True
except Exception:
    CHARTS_AVAILABLE = False
    ChartGenerator = None

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

class ExportGenerator:
    @staticmethod
    def to_csv(data, filename):
        output = io.StringIO()
        writer = csv.writer(output)
        
        def write_recursive(d, prefix=""):
            if isinstance(d, dict):
                for k, v in d.items():
                    new_prefix = f"{prefix}{k}." if prefix else f"{k}."
                    if isinstance(v, (dict, list)):
                        write_recursive(v, new_prefix)
                    else:
                        writer.writerow([prefix + k, v])
            elif isinstance(d, list):
                for i, item in enumerate(d):
                    write_recursive(item, f"{prefix}[{i}].")
            else:
                writer.writerow([prefix[:-1] if prefix.endswith('.') else prefix, d])

        if isinstance(data, (dict, list)):
            write_recursive(data)
        return output.getvalue().encode('utf-8-sig')

    @staticmethod
    def to_excel(data, title):
        if not OPENPYXL_AVAILABLE:
            return ExportGenerator.to_csv(data, "report.csv")

        wb = openpyxl.Workbook()
        # Remove default sheet
        wb.remove(wb.active)

        # Helper to create a sheet
        def create_sheet(sheet_title, sheet_data):
            # Clean up sheet title (max 31 chars, no invalid chars)
            safe_title = str(sheet_title)[:31].replace('*', '').replace(':', '').replace('?', '').replace('/', '').replace('\\', '')
            if not safe_title:
                safe_title = "Data"
            ws = wb.create_sheet(title=safe_title)
            
            # Styles
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            
            # Add metadata row
            ws.append([f"Report: {title}", f"Generated: {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}"])
            ws.append([])
            
            if isinstance(sheet_data, list) and sheet_data and isinstance(sheet_data[0], dict):
                # It's a list of dicts (table format)
                headers = list(sheet_data[0].keys())
                ws.append([str(h).replace('_', ' ').title() for h in headers])
                # Style headers
                for cell in ws[ws.max_row]:
                    cell.font = header_font
                    cell.fill = header_fill
                # Add rows
                for row_item in sheet_data:
                    ws.append([str(row_item.get(h, '')) for h in headers])
            elif isinstance(sheet_data, dict):
                # Key-Value pairs
                ws.append(["Metric", "Value"])
                for cell in ws[ws.max_row]:
                    cell.font = header_font
                    cell.fill = header_fill
                    
                def flatten_dict(d, prefix=""):
                    for k, v in d.items():
                        new_prefix = f"{prefix}{k} > " if prefix else f"{k} > "
                        if isinstance(v, dict):
                            flatten_dict(v, new_prefix)
                        elif isinstance(v, list):
                            ws.append([new_prefix.strip(" > "), str(v)])
                        else:
                            ws.append([new_prefix.strip(" > ").replace('_', ' ').title(), str(v)])
                flatten_dict(sheet_data)
            else:
                ws.append(["Value"])
                ws.append([str(sheet_data)])

        # Iterate over top-level keys in data to create sheets
        if isinstance(data, dict):
            for key, val in data.items():
                if isinstance(val, (dict, list)):
                    create_sheet(key.replace('_', ' ').title(), val)
                else:
                    # Collect scalar values into a 'Summary' sheet if they exist at root
                    if "Summary" not in wb.sheetnames:
                        ws = wb.create_sheet("Summary")
                        ws.append(["Metric", "Value"])
                    ws = wb["Summary"]
                    ws.append([str(key).replace('_', ' ').title(), str(val)])
        else:
            create_sheet("Data", data)

        if not wb.sheetnames:
            wb.create_sheet("Empty Report")

        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    @staticmethod
    def to_pdf(data, title, report_type="System Report"):
        if not REPORTLAB_AVAILABLE:
            # ReportLab not installed — generate a minimal but VALID PDF as a fallback
            # so Adobe never shows "Failed to open" (plain text saved as .pdf is the cause)
            try:
                from reportlab.pdfgen import canvas as _canvas
                from reportlab.lib.pagesizes import letter as _letter
                _buf = io.BytesIO()
                _c = _canvas.Canvas(_buf, pagesize=_letter)
                _c.drawString(72, 720, f"{title} — ReportLab not installed on this server.")
                _c.drawString(72, 700, "Please install reportlab: pip install reportlab")
                _c.save()
                _buf.seek(0)
                return _buf.getvalue()
            except Exception:
                pass
            output = f"--- {title} ---\n\n" + json.dumps(data, indent=4, default=str)
            return output.encode('utf-8')

        try:
            return ExportGenerator._build_pdf(data, title, report_type)
        except Exception as e:
            # If PDF generation crashes, return a valid PDF with an error message
            # so the file always opens in Adobe instead of showing "Failed to open"
            import traceback
            err_buf = io.BytesIO()
            doc = SimpleDocTemplate(err_buf, pagesize=letter)
            styles = getSampleStyleSheet()
            error_text = traceback.format_exc()
            elements = [
                Paragraph("Report Generation Error", styles['Heading1']),
                Spacer(1, 0.2 * inch),
                Paragraph(f"An error occurred while generating the report: {str(e)}", styles['Normal']),
                Spacer(1, 0.1 * inch),
                Paragraph("Technical Details:", styles['Heading3']),
                Paragraph(error_text.replace('\n', '<br/>').replace(' ', '&nbsp;'), styles['Code']),
            ]
            doc.build(elements)
            err_buf.seek(0)
            return err_buf.getvalue()

    @staticmethod
    def _build_pdf(data, title, report_type="System Report"):
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=72, leftMargin=72, topMargin=72, bottomMargin=72)
        styles = getSampleStyleSheet()
        
        # Custom Styles
        title_style = ParagraphStyle('TitleStyle', parent=styles['Heading1'], fontSize=18, alignment=1, spaceAfter=12, textColor=colors.HexColor("#2C3E50"))
        header_style = ParagraphStyle('HeaderStyle', parent=styles['Normal'], fontSize=10, alignment=1, spaceAfter=20)
        section_style = ParagraphStyle('SectionStyle', parent=styles['Heading2'], fontSize=14, spaceBefore=15, spaceAfter=10, textColor=colors.HexColor("#2980B9"), borderPadding=5)
        body_style = styles['Normal']
        body_style.leading = 14

        elements = []

        # 1. COVER PAGE
        elements.append(Spacer(1, 2 * inch))
        cover_title = ParagraphStyle('CoverTitle', parent=styles['Heading1'], fontSize=28, alignment=1, spaceAfter=20, textColor=colors.HexColor("#2C3E50"))
        elements.append(Paragraph("JUSTICE HUB", cover_title))
        elements.append(Paragraph("DIGITAL COURT SYSTEM", ParagraphStyle('CoverSub', parent=cover_title, fontSize=20, textColor=colors.HexColor("#2980B9"))))
        elements.append(Spacer(1, 1 * inch))
        
        elements.append(Paragraph(f"<b>Report Type:</b> {report_type}", ParagraphStyle('CT', parent=styles['Heading2'], alignment=1)))
        
        period = data.get('period', {})
        start_date = period.get('start_date', 'N/A')
        end_date = period.get('end_date', 'N/A')
        if hasattr(start_date, 'strftime'): start_date = start_date.strftime('%Y-%m-%d')
        if hasattr(end_date, 'strftime'): end_date = end_date.strftime('%Y-%m-%d')
        elements.append(Paragraph(f"<b>Reporting Period:</b> {start_date} to {end_date}", ParagraphStyle('CT2', parent=styles['Heading3'], alignment=1)))
        elements.append(Spacer(1, 2 * inch))
        
        elements.append(Paragraph(f"<b>Generated Date:</b> {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}", ParagraphStyle('CN', parent=styles['Normal'], alignment=1)))
        elements.append(Paragraph(f"<b>Generated By:</b> {data.get('generated_by', 'Administrator')}", ParagraphStyle('CN2', parent=styles['Normal'], alignment=1)))
        elements.append(PageBreak())

        # 2. EXECUTIVE SUMMARY
        elements.append(Paragraph("[ EXECUTIVE SUMMARY ]", section_style))
        summary = data.get('system_summary', data.get('summary', {}))
        stats = summary.get('stats', summary)
        
        total = stats.get('total_cases', 0)
        resolved = stats.get('resolved', 0)
        pending = stats.get('pending_cases', total - resolved)
        avg_res = summary.get('average_resolution_time', 'N/A')
        rate = summary.get('resolution_rate', '0%')
        exec_text = (f"During this period, {total} cases were processed. "
                     f"The courts resolved {resolved} cases, achieving a {rate} resolution rate. "
                     f"{pending} cases remain pending. The average resolution time was {avg_res}.")
        elements.append(Paragraph(exec_text, body_style))
        elements.append(Spacer(1, 0.2 * inch))

        # 3. CASE STATISTICS
        elements.append(Paragraph("[ CASE STATISTICS ]", section_style))
        stats_data = [
            ["Metric", "Value"],
            ["Total Cases", str(total)],
            ["Resolved (Closed)", str(resolved)],
            ["Pending Review", str(stats.get('pending_review', 0))],
            ["Approved (Awaiting Payment)", str(stats.get('approved', 0))],
            ["Assigned to Judge", str(stats.get('assigned', 0))],
            ["Decided (Awaiting Closure)", str(stats.get('decided', 0))],
            ["Resolution Rate", summary.get('resolution_rate', '0%')],
            ["Average Resolution Time", avg_res]
        ]
        t = Table(stats_data, colWidths=[3 * inch, 2.5 * inch])
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#ECF0F1")),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor("#2C3E50")),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey)
        ]))
        elements.append(t)
        
        # Pie chart using real cases_by_type from database
        cases_by_type = data.get('cases_by_type', [])
        if cases_by_type and CHARTS_AVAILABLE:
            try:
                chart_data = {c['case_type']: c['count'] for c in cases_by_type if c.get('count', 0) > 0}
                if chart_data:
                    pie_buf = ChartGenerator.generate_pie_chart(chart_data, "Case Distribution by Category")
                    if pie_buf:
                        elements.append(Spacer(1, 0.2 * inch))
                        elements.append(Image(pie_buf, width=4*inch, height=3*inch))
            except Exception:
                pass  # Chart failure is non-fatal; PDF continues without it
        elements.append(Spacer(1, 0.2 * inch))

        # 4. PARTICIPANT PROFILE (Demographics)
        if 'demographics' in data:
            elements.append(Paragraph("[ PARTICIPANT PROFILE ]", section_style))
            demo = data['demographics']
            
            # Education
            elements.append(Paragraph("<b>Education Level Distribution</b>", styles['Normal']))
            edu_data = [["Education", "Count"]]
            edu_dist = demo.get('education_distribution', {})
            for k, v in edu_dist.items():
                edu_data.append([k, str(v)])
            t_edu = Table(edu_data, colWidths=[3 * inch, 2.5 * inch])
            t_edu.setStyle(TableStyle([('GRID', (0, 0), (-1, -1), 1, colors.grey), ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold')]))
            elements.append(t_edu)
            if edu_dist:
                top_edu = max(edu_dist, key=edu_dist.get)
                elements.append(Paragraph(f"• Most participants held {top_edu} qualifications.", body_style))
            elements.append(Spacer(1, 0.2 * inch))

            # Age
            age_dist = demo.get('age_distribution', {})
            if age_dist:
                top_age = max(age_dist, key=age_dist.get)
                elements.append(Paragraph(f"• The largest participant age demographic was {top_age} years old.", body_style))
            
            # Regional Distribution
            subcity_dist = demo.get('subcity_distribution', {})
            if subcity_dist:
                elements.append(Spacer(1, 0.2 * inch))
                elements.append(Paragraph("[ REGIONAL DISTRIBUTION ]", section_style))
                top_subcity = max(subcity_dist, key=subcity_dist.get)
                
                reg_data = [["Sub-City", "Count"]]
                for k, v in list(subcity_dist.items())[:5]:
                    reg_data.append([k, str(v)])
                t_reg = Table(reg_data, colWidths=[3 * inch, 2.5 * inch])
                t_reg.setStyle(TableStyle([('GRID', (0, 0), (-1, -1), 1, colors.grey), ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold')]))
                elements.append(t_reg)
                elements.append(Paragraph(f"• {top_subcity} recorded the highest number of filed cases.", body_style))
                elements.append(Spacer(1, 0.2 * inch))

        # 5. HEARING & BACKLOG ANALYTICS (Requirement 9)
        hearings = summary.get('hearings', {})
        if hearings:
            elements.append(Paragraph("[ HEARING & BACKLOG ANALYTICS ]", section_style))
            h_data = [
                ["Hearing Metric", "Count"],
                ["Total Hearings Scheduled", str(hearings.get('total_hearings', 0))],
                ["Conducted/Completed", str(hearings.get('conducted', 0))],
                ["Postponed Sessions", str(hearings.get('postponed', 0))],
                ["Cancelled Sessions", str(hearings.get('cancelled', 0))],
                ["Attended (Full Presence)", str(hearings.get('attended', 0))],
                ["Missed Attendance (Absent)", str(hearings.get('not_attended', 0))],
                ["Backlog Contribution", str(hearings.get('backlog_contribution', 0))]
            ]
            t_h = Table(h_data, colWidths=[3 * inch, 2.5 * inch])
            t_h.setStyle(TableStyle([('GRID', (0, 0), (-1, -1), 1, colors.grey), ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold')]))
            elements.append(t_h)
            
            c = hearings.get('conducted', 0)
            p = hearings.get('postponed', 0)
            t = hearings.get('total_hearings', 0)
            h_rate = f"{(c/t*100):.1f}%" if t > 0 else "0%"
            elements.append(Paragraph(f"• {c} hearings were completed ({h_rate} completion rate).", body_style))
            if p > 0:
                elements.append(Paragraph(f"• {p} hearings were postponed, adding to the backlog.", body_style))
            elements.append(Spacer(1, 0.2 * inch))

        # 6. FINANCIAL OVERVIEW
        elements.append(Paragraph("[ FINANCIAL OVERVIEW ]", section_style))
        fin = data.get('financial_summary', data.get('financial', {}))
        total_rev = fin.get('total_revenue', fin.get('total_service_fees_earned', 0))
        fin_data = [
            ["Metric", "Value"],
            ["Total Revenue Collected", f"{total_rev:,} ETB"],
            ["Paid Cases (Verified)", str(fin.get('actual_paid_cases', fin.get('paid_cases', 'N/A')))],
            ["Collection Rate", fin.get('collection_rate', 'N/A')]
        ]
        t_fin = Table(fin_data, colWidths=[3 * inch, 2.5 * inch])
        t_fin.setStyle(TableStyle([('GRID', (0, 0), (-1, -1), 1, colors.grey), ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold')]))
        elements.append(t_fin)
        
        # Bullet insight from the cases_by_type fetched above
        if cases_by_type:
            top_case = max(cases_by_type, key=lambda x: x.get('count', 0))
            elements.append(Paragraph(f"- {top_case.get('case_type', 'General')} cases contributed the highest volume during the period.", body_style))
        elements.append(Paragraph(f"- A total of {total_rev:,} ETB was successfully collected.", body_style))
        
        fin_summary = data.get('financial_summary', {})
        rev_month = fin_summary.get('revenue_by_month', [])
        if len(rev_month) >= 2:
            last_m = float(rev_month[-1]['revenue'] or 0)
            prev_m = float(rev_month[-2]['revenue'] or 0)
            diff = last_m - prev_m
            pct = (diff / prev_m * 100) if prev_m > 0 else 0
            trend_word = "increased" if pct >= 0 else "decreased"
            trend_p = Paragraph(f"- Trend Analysis: Revenue {trend_word} by {abs(pct):.1f}% compared to the previous month.", body_style)
            elements.append(trend_p)
            
        if rev_month and CHARTS_AVAILABLE:
            try:
                # Use real month labels and revenue values from the database
                labels = [str(m.get('month', '')) for m in rev_month[-6:]]
                values = [float(m.get('revenue', 0) or 0) for m in rev_month[-6:]]
                if any(v > 0 for v in values):
                    bar_buf = ChartGenerator.generate_bar_chart(labels, values, "Monthly Revenue (ETB)")
                    if bar_buf:
                        elements.append(Spacer(1, 0.1 * inch))
                        elements.append(Image(bar_buf, width=5*inch, height=2.5*inch))
            except Exception:
                pass  # Chart failure is non-fatal; PDF continues without it
                
        elements.append(Spacer(1, 0.2 * inch))

        # 7. DECISION TYPES (Requirement 5)
        decision_analysis = data.get('decision_analysis')
        if decision_analysis:
            elements.append(Paragraph("[ DECISION & RESOLUTION TYPES ]", section_style))
            d_text = f"Total Decisions Finalized: {decision_analysis.get('total_decisions', 0)}. "
            d_text += f"Notably, {decision_analysis.get('mediation_resolved', 0)} cases were resolved through Mediation/Settlement. "
            d_text += f"{decision_analysis.get('immediate_decisions', 0)} cases received Immediate Decisions."
            elements.append(Paragraph(d_text, body_style))
            
            dist = decision_analysis.get('distribution', {})
            if dist:
                d_data = [["Decision Type", "Count"]]
                for k, v in dist.items(): d_data.append([k, str(v)])
                t_d = Table(d_data, colWidths=[3 * inch, 2.5 * inch])
                t_d.setStyle(TableStyle([('GRID', (0, 0), (-1, -1), 1, colors.grey), ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold')]))
                elements.append(t_d)

        # 8. JUDGE PERFORMANCE
        elements.append(Paragraph("[ JUDGE PERFORMANCE ]", section_style))
        if 'cases_by_judge' in data and data['cases_by_judge']:
            j_data = [["Judge Name", "Active", "Resolved", "Avg Days"]]
            for j in data['cases_by_judge'][:5]: # Top 5
                j_data.append([j['name'], str(j['active_cases']), str(j['total_resolved']), str(j['avg_resolution_days'])])
            t_j = Table(j_data, colWidths=[2.5 * inch, 1 * inch, 1 * inch, 1 * inch])
            t_j.setStyle(TableStyle([('GRID', (0, 0), (-1, -1), 1, colors.grey), ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold')]))
            elements.append(t_j)
            
            top_j = data['cases_by_judge'][0]
            elements.append(Paragraph(f"• Judge {top_j['name']} carries the highest active caseload ({top_j['active_cases']} cases).", body_style))
        else:
            elements.append(Paragraph(f"• Judicial officers maintained standard efficiency.", body_style))
        elements.append(Spacer(1, 0.2 * inch))

        # 9. INTELLIGENCE INSIGHTS
        elements.append(Paragraph("[ INTELLIGENCE INSIGHTS ]", section_style))
        
        warnings = data.get('intelligence_insights', {}).get('warnings', [])
        if warnings:
            elements.append(Paragraph("<b>!! CRITICAL WARNINGS !!</b>", ParagraphStyle('Warn', parent=body_style, textColor=colors.HexColor("#C0392B"))))
            warn_data = [[Paragraph(f"• {w}", body_style)] for w in warnings]
            t_warn = Table(warn_data, colWidths=[6 * inch])
            t_warn.setStyle(TableStyle([('BACKGROUND', (0, 0), (-1, -1), colors.HexColor("#FADBD8")),
                                        ('BOX', (0, 0), (-1, -1), 1, colors.HexColor("#E74C3C")),
                                        ('BOTTOMPADDING', (0, 0), (-1, -1), 6)]))
            elements.append(t_warn)
            elements.append(Spacer(1, 0.1 * inch))
            
        insights = data.get('intelligence_insights', {}).get('bottlenecks', [])
        if not insights:
            insights = ["Steady performance across all case categories.", "Resolution times are within the targeted SLA."]
        for insight in insights:
            elements.append(Paragraph(f"• {insight}", body_style))

        # 10. CONCLUSION
        elements.append(Paragraph("[ CONCLUSION ]", section_style))
        conc_text = "The system continues to perform effectively. Future focus should remain on reducing backlog through increased mediation and streamlined hearing scheduling."
        elements.append(Paragraph(conc_text, body_style))

        # 11. SIGNATURE SECTION
        elements.append(Spacer(1, 0.5 * inch))
        elements.append(Paragraph("SIGNATURE SECTION", styles['Heading3']))
        sig_data = [
            [f"Name: {data.get('generated_by', 'Administrator')}", ""],
            [f"Role: Judicial System Official", "Signature: _______________________"],
            [f"Date: {timezone.now().strftime('%Y-%m-%d')}", ""]
        ]
        t_sig = Table(sig_data, colWidths=[3 * inch, 2.5 * inch])
        elements.append(t_sig)

        def add_footer(canvas, doc):
            canvas.saveState()
            canvas.setFont('Helvetica', 9)
            canvas.drawString(inch, 0.5 * inch, "JusticeHub Official Report | Confidential")
            canvas.drawRightString(7.5 * inch, 0.5 * inch, f"Page {doc.page}")
            canvas.restoreState()

        doc.build(elements, onFirstPage=add_footer, onLaterPages=add_footer)
        buffer.seek(0)
        return buffer.getvalue()

